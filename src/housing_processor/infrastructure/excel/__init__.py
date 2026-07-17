from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from housing_processor.application.contracts.excel import HousingWorkbookProjection


class OpenpyxlExcelRenderer:
    """Phase 1 stub: writes header row and projection rows with bold contact names."""

    HEADERS = [
        "group_number",
        "applicant_name",
        "is_contact",
        "phone",
        "email",
        "gpa",
        "requested_properties",
        "expected_group_size",
        "application_received_date",
        "group_status",
        "review_notes",
    ]

    def render(self, projection: HousingWorkbookProjection) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Applicants"
        sheet.append(self.HEADERS)
        sheet.freeze_panes = "A2"

        bold_font = Font(bold=True)
        for row in projection.rows:
            sheet.append(
                [
                    row.group_number,
                    row.applicant_name,
                    row.is_contact,
                    row.phone,
                    row.email,
                    str(row.gpa) if row.gpa is not None else None,
                    ", ".join(row.requested_properties),
                    row.expected_group_size,
                    row.application_received_date.isoformat()
                    if row.application_received_date
                    else None,
                    row.group_status,
                    row.review_notes,
                ]
            )
            if row.is_contact:
                sheet.cell(row=sheet.max_row, column=2).font = bold_font

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
